import openpyxl

# --- 1. Create a new workbook ---
workbook = openpyxl.Workbook()

# Rename the default sheet to "Grades"
sheet = workbook.active
sheet.title = "Grades"

# --- 2. Given data ---
data = {
    "Joe": {
        "math": 65,
        "science": 78,
        "english": 98,
        "gym": 89
    },
    "Bill": {
        "math": 55,
        "science": 72,
        "english": 87,
        "gym": 95
    },
    "Tim": {
        "math": 100,
        "science": 45,
        "english": 75,
        "gym": 92
    },
    "Sally": {
        "math": 30,
        "science": 25,
        "english": 45,
        "gym": 100
    },
    "Jane": {
        "math": 100,
        "science": 100,
        "english": 100,
        "gym": 60
    }
}

# --- 3. Write headers ---
headers = ["Name", "Math", "Science", "English", "Gym"]

for col_index, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=col_index).value = header

# --- 4. Populate sheet with student data ---
row = 2
for student, subjects in data.items():
    sheet.cell(row=row, column=1).value = student
    sheet.cell(row=row, column=2).value = subjects["math"]
    sheet.cell(row=row, column=3).value = subjects["science"]
    sheet.cell(row=row, column=4).value = subjects["english"]
    sheet.cell(row=row, column=5).value = subjects["gym"]
    row += 1

# --- 5. Save the workbook ---
workbook.save("grades.xlsx")

print("grades.xlsx created successfully!")

# Exercise 2

from openpyxl.styles import Font, PatternFill

# --- 1. Create a new workbook ---
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Grades"

# --- 2. Given data ---

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

for col_index, header in enumerate(headers, start=1):
    cell = sheet.cell(row=1, column=col_index)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill

# --- 4. Add a row for each student's grades ---
row = 2
for student, subjects in data.items():
    sheet.cell(row=row, column=1).value = student
    sheet.cell(row=row, column=2).value = subjects["math"]
    sheet.cell(row=row, column=3).value = subjects["science"]
    sheet.cell(row=row, column=4).value = subjects["english"]
    sheet.cell(row=row, column=5).value = subjects["gym"]
    row += 1

# --- 5. Add average formulas in the last row ---
average_row = row
for col in range(2, 6):  # Columns B to E (Math to Gym)
    col_letter = sheet.cell(row=1, column=col).column_letter
    formula = f"=AVERAGE({col_letter}2:{col_letter}{row-1})"
    sheet.cell(row=average_row, column=col).value = formula

sheet.cell(row=average_row, column=1).value = "Average"

# --- 6. Save workbook ---
workbook.save("grades.xlsx")
print("grades.xlsx created with averages and formatted header!")
