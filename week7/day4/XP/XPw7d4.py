import openpyxl
import os
workbook_variable = openpyxl.Workbook()
sheet = workbook_variable.active
sheet.title = "MySheet"
sheet['A1'] = "Hello"
sheet['B1'] = 123


# Create workbook
wb = openpyxl.Workbook()

# Define path dynamically
project_folder = os.path.dirname(__file__)  # folder where script.py is located
file_path = os.path.join(project_folder, "my_new_workbook.xlsx")

# Save workbook
wb.save(file_path)

# Refer to the current (active) worksheet
worksheet_variable = workbook_variable.active
print(worksheet_variable.title)


worksheet_variable.title = "MyNewSheet"


row_number = 2
column_number = 5
worksheet_variable.cell(row=row_number, column=column_number).value = 25
workbook_variable.save("my_new_workbook.xlsx")
workbook_variable.close()


# Exercise 2

# --- 1. Open the workbook containing the list of plants ---
# workbook = openpyxl.load_workbook.file_path("plants.xlsx")
file_path = os.path.join(project_folder, "plants.xlsx")
workbook = openpyxl.load_workbook(file_path)


# --- 2. Get a reference to the first sheet (called Sheet1) ---
sheet = workbook["Sheet1"]     # or workbook.active if Sheet1 is first

# --- 3. Start at cell A2 ---
cell = sheet["A2"]

# --- 4. Loop until a blank cell is encountered ---
while cell.value is not None:

    # Cell offset: 0 rows down, 7 columns to the right (i.e., column H)
    stock_cell = cell.offset(row=0, column=7)

    # If this reads "No", print plant name
    if str(stock_cell.value).strip().lower() == "no":
        print(cell.value)

    # Move down one row (A3 → A4 → A5...)
    cell = cell.offset(row=1, column=0)

# Program ends safely when a blank plant name is found
print("Finished scanning.")

workbook.close()

# Exercise 3

import pandas as pd

# --- 1. Load Excel File (give full path if needed) ---
# Load using pandas
file_path = os.path.join(project_folder, "data.xlsx")
workbook = openpyxl.load_workbook(file_path)
df = pd.read_excel(file_path)

# --- 2. Filter rows where Sales > 1000 ---
filtered_df = df[df['Sales'] > 1000]

print("Filtered Data:")
print(filtered_df)

# --- 3. Write filtered data back to the Excel file using openpyxl ---

# Load workbook using openpyxl
wb = openpyxl.load_workbook(file_path)

# Create (or overwrite) a sheet for filtered data
if "FilteredData" in wb.sheetnames:
    ws = wb["FilteredData"]
    wb.remove(ws)        # remove old version to avoid duplicates

ws = wb.create_sheet("FilteredData")

# Write header
for col_index, column_name in enumerate(filtered_df.columns, start=1):
    ws.cell(row=1, column=col_index).value = column_name

# Write rows
for row_index, row_data in filtered_df.iterrows():
    for col_index, value in enumerate(row_data, start=1):
        ws.cell(row=row_index + 2, column=col_index).value = value

# Save workbook
wb.save(file_path)

print("Filtered data written to sheet 'FilteredData' in data.xlsx")


# Exercise 4

import matplotlib.pyplot as plt

# --- 1. Read data from Excel file ---
file_path = os.path.join(project_folder, "productSales.xlsx")
workbook = openpyxl.load_workbook(file_path)
df = pd.read_excel(file_path)


# --- 2. Group by product and sum sales ---
sales_summary = df.groupby("product")["sales"].sum().reset_index()

print("Sales Summary:")
print(sales_summary)

# --- 3. Export summary to a new Excel file ---
with pd.ExcelWriter("sales_report.xlsx", engine="openpyxl") as writer:
    sales_summary.to_excel(writer, sheet_name="Summary", index=False)
    # writer.save() is automatically handled inside the context manager
    # writer.close() is also automatically handled


print("sales_report.xlsx created successfully!")
